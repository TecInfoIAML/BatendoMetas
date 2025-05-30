# Importa os módulos necessários do Flask
import flask import Flask, render_template, request, redirect

# Importa funções da biblioteca openpyxl para criar e manipular arquivos Excel
from openpyxl import Workbook, load_workbook

# Importa a biblioteca OS para verificar a existência do arquivo Excel
import os

# Cria uma aplicação Flask
app = Flask(__name__)

# Define o nome do arquivo Excel onde os dados serão salvos
ARQUIVO = 'vendas.xlsx'

# Verifica se o arquivo Excel já existe. Se não existir, cria um novo com cabeçalho.
if not os.path.exists(ARQUIVO):
    wb = Workbook()  # Cria um novo arquivo Excel
    ws = wb.active   # Seleciona a planilha ativa (por padrão é chamada 'Sheet')
    ws.append(["Nome", "Vendas", "Meta"])  # Adiciona uma linha de cabeçalho
    wb.save(ARQUIVO)  # Salva o arquivo com o nome definido

# Rota principal do site (formulário de cadastro de vendas)
@app.route('/')
def index():
    return render_template('index.html')  # Retorna o conteúdo do arquivo index.html

# Rota que processa os dados do formulário e salva no Excel
@app.route('/salvar', methods=['POST'])
def salvar():
    # Captura os dados enviados pelo formulário HTML
    nome = request.form['nome']  # Nome do funcionário
    vendas = float(request.form['vendas'])  # Total de vendas, convertido para número decimal
    meta = float(request.form['meta'])  # Meta de vendas, também convertida para número decimal

    # Abre o arquivo Excel existente e seleciona a planilha ativa
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    # Adiciona uma nova linha com os dados informados no formulário
    ws.append([nome, vendas, meta])

    # Salva o arquivo Excel com os novos dados
    wb.save(ARQUIVO)

    # Redireciona para a rota /analisar passando o nome do funcionário na URL
    return redirect('/analisar?nome=' + nome)

# Rota que analisa se o funcionário bateu a meta e calcula o bônus
@app.route('/analisar')
def analisar():
    # Pega o nome do funcionário enviado como parâmetro pela URL
    nome_param = request.args.get('nome')

    # Abre o arquivo Excel e acessa a planilha ativa
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    # Percorre todas as linhas da planilha (a partir da segunda, pois a primeira é o cabeçalho) values _only retorna apenas os valores 
    for linha in ws.iter_rows(min_row=2, values_only=True):
        nome, vendas, meta = linha  # Desempacota os valores da linha

        # Verifica se o nome da linha atual é o mesmo nome enviado na URL
        if nome == nome_param:
            meta_batida = vendas >= meta  # Verifica se a meta foi atingida
            bonus = round(vendas * 0.15, 2) if meta_batida else 0  # Calcula o bônus (15%) se a meta for batida

            # Renderiza a página resultado.html passando os dados calculados
            return render_template('resultado.html', nome=nome, meta_batida=meta_batida, bonus=bonus)

    # Caso o nome não seja encontrado na planilha
    return "Funcionário não encontrado."

# Rota que mostra o histórico de todos os funcionários cadastrados
@app.route('/historico')
def historico():
    # Abre o arquivo Excel e acessa a planilha ativa
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    # Converte os dados da planilha (a partir da linha 2) em uma lista de tuplas
    dados = list(ws.iter_rows(min_row=2, values_only=True))

    # Renderiza o template historico.html passando os dados da planilha
    return render_template('historico.html', dados=dados)

# Inicia o servidor Flask em modo debug (útil para testes e desenvolvimento)
if __name__ == '__main__':
    app.run(debug=True)
